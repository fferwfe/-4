import streamlit as st
import pandas as pd
import io
import re
import os
from google.cloud import vision
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

# 設定網頁標題與圖示
st.set_page_config(page_title="學界二班團購助手", layout="wide")

# --- 1. 初始化 Google AI (從 Secrets 讀取) ---
def init_vision():
    # 建議在 Streamlit Secrets 設定金鑰，避免 key.json 外流
    if "gcp_service_account" in st.secrets:
        import json
        with open("key.json", "w") as f:
            json.dump(dict(st.secrets["gcp_service_account"]), f)
        os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'key.json'

# --- 2. 商品設定介面 ---
st.title("🛍️ 團購截圖自動轉 Excel 系統")
st.markdown("填寫本週商品資訊，上傳 LINE 截圖，即可下載格式化 Excel。")

with st.expander("📝 第一步：設定本週商品資訊", expanded=True):
    prod_df = pd.DataFrame([
        {"代碼": "A", "品名": "航空米果", "單價": 150, "單位": "顆"},
        {"代碼": "B", "品名": "餅乾", "單價": 220, "單位": "包"},
        {"代碼": "C", "品名": "飲料", "單價": 170, "單位": "罐"}
    ])
    edited_df = st.data_editor(prod_df, num_rows="dynamic")
    config = edited_df.set_index("代碼").to_dict('index')
    default_item = edited_df.iloc[0] # 預設抓第一行

# --- 3. 圖片上傳區 ---
st.subheader("📸 第二步：上傳留言截圖")
uploaded_files = st.file_uploader("可一次選擇多張截圖", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True)

if uploaded_files:
    all_data = []
    
    # 模擬/實際 AI 辨識邏輯
    for file in uploaded_files:
        # 這裡會放入 Vision AI 辨識代碼
        # 範例邏輯：抓取像「人名 +1」的格式
        # 目前先以您提供的截圖內容做範例預覽
        st.success(f"已讀取圖片: {file.name}")
    
    # 模擬辨識結果 (這部分會由 AI 自動產生)
    all_data = [
        {"姓名": "胡珍華", "數量": 1}, {"姓名": "陳昱佑", "數量": 1},
        {"姓名": "雅瑜", "數量": 1}, {"姓名": "淑妹", "數量": 1},
        {"姓名": "詩茹", "數量": 1}, {"姓名": "陳敬岳", "數量": 1},
        {"姓名": "何婕瑀", "數量": 1}, {"姓名": "陳政男", "數量": 1},
        {"姓名": "胡雋", "數量": 1}
    ]
    
    st.write("📋 辨識清單預覽：", pd.DataFrame(all_data))

    # --- 4. 生成 Excel (精準還原您的航空米果格式) ---
    if st.button("🚀 生成 2025 標準格式 Excel"):
        output = io.BytesIO()
        wb = Workbook()
        
        # --- 分頁一：付款單 (橫向格式) ---
        ws1 = wb.active
        ws1.title = "付款單"
        title = f"學 界 二 班   {default_item['品名']}"
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_data))
        ws1['A1'] = title
        
        for i, order in enumerate(all_data, 1):
            ws1.cell(row=2, column=i, value=f"學二  {default_item['品名']}")
            ws1.cell(row=3, column=i, value="N1")
            ws1.cell(row=4, column=i, value=order['姓名'])
            ws1.cell(row=5, column=i, value=order['數量'])
            ws1.cell(row=6, column=i, value=default_item['單位'])
            ws1.cell(row=7, column=i, value=default_item['單價'])
            ws1.cell(row=8, column=i, value="元")

        # --- 分頁二：對帳單 (縱向格式) ---
        ws2 = wb.create_sheet("對帳單")
        ws2['A1'] = title
        ws2['C2'] = "應付款項"
        ws2['D2'] = "付款狀態"
        ws2['A3'] = "一個"
        ws2['B3'] = default_item['單價']
        
        total = 0
        for r, order in enumerate(all_data, 5):
            ws2.cell(row=r, column=1, value=order['姓名'])
            ws2.cell(row=r, column=2, value=order['數量'])
            ws2.cell(row=r, column=3, value=order['數量'] * default_item['單價'])
            total += order['數量'] * default_item['單價']
        
        ws2.cell(row=len(all_data)+6, column=1, value="總計")
        ws2.cell(row=len(all_data)+6, column=3, value=total)

        # --- 分頁三：商品標籤 ---
        ws3 = wb.create_sheet("商品標籤")
        for i, order in enumerate(all_data):
            row_idx = i * 2 + 1
            ws3.cell(row=row_idx, column=1, value=f"學二{default_item['品名']}")
            ws3.cell(row=row_idx+1, column=1, value=order['姓名'])
            ws3.cell(row=row_idx+1, column=2, value=order['數量'])

        wb.save(output)
        st.download_button(
            label="⬇️ 點我下載 Excel",
            data=output.getvalue(),
            file_name=f"2025付款單_{default_item['品名']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
